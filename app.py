import os, json
from flask import Flask, request, jsonify, send_from_directory, redirect, session
import openpyxl
from collections import defaultdict

app = Flask(__name__, static_folder='static')
app.secret_key = os.environ.get('SECRET_KEY', 'anb-secret-2024')
ADMIN_PASSWORD = os.environ.get('ADMIN_PASSWORD', 'anb2024')
DATA_FILE = 'data/anb_data.json'
UPLOAD_FOLDER = 'uploads'
os.makedirs('data', exist_ok=True)
os.makedirs(UPLOAD_FOLDER, exist_ok=True)

FLAG_MAP = {
    'CHINA':'CN','HONG KONG':'HK','MALAYSIA':'MY','PHILIPPINES':'PH',
    'VIETNAM':'VN','THAILAND':'TH','SINGAPORE':'SG','LAOS':'LA',
    'UNITED STATES':'US','MEXICO':'MX','PERU':'PE','CHILE':'CL',
    'BRAZIL':'BR','ECUADOR':'EC','GERMANY':'DE','PORTUGAL':'PT',
    'SPAIN':'ES','GREECE':'GR','SWITZERLAND':'CH','ALBANIA':'AL',
    'ISRAEL':'IL','EGYPT':'EG','UNITED ARAB EMIRATES':'AE','UNITED ARAB EMIRAT':'AE'
}
POOL_FLAG = {'China':'CN','MALAYSIA':'MY','PHILIPPINES':'PH','PERU':'PE'}
SOLAPAS = ['ASADO','PECHO','R&L HILTON','R&L W','RUEDA']
SKIP = {'Stock de cuartos','Produccion','Rechazos','Degradado a MB1','Pasado a MB1'}
POOL_IDS = {'China','MALAYSIA','PHILIPPINES','PERU','SALDO'}


def process_excel(path):
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)

    ws_dv = wb['DETALLE VENTAS']
    rows_dv = list(ws_dv.iter_rows(values_only=True))
    gen_kgu = defaultdict(list)
    for row in rows_dv[3:]:
        gen = str(row[10]) if row[10] else ''
        kgu = row[19]; kgv = row[16]
        if gen and gen != 'None' and isinstance(kgu,(int,float)) and kgu>0 and isinstance(kgv,(int,float)) and kgv>0:
            gen_kgu[gen].append((kgu, kgv))
    KGU = {gen: round(sum(v[0]*v[1] for v in vals)/sum(v[1] for v in vals), 3) for gen, vals in gen_kgu.items()}

    venta_solapa = {}
    sol_out = {}

    for solapa in SOLAPAS:
        if solapa not in wb.sheetnames:
            continue
        ws = wb[solapa]
        rows = list(ws.iter_rows(values_only=True))
        header = rows[3]
        is_rlw = solapa == 'R&L W'

        genericos_mb1 = []; idx_mb1 = []
        for i, h in enumerate(header[10:], start=10):
            if not h: continue
            if str(h) in ('INSTRUCCIONES','ALERTAS','Sem','Venta','Concepto'): break
            genericos_mb1.append(str(h)); idx_mb1.append(i)

        genericos_mb2 = []; idx_mb2 = []
        if is_rlw:
            for i, h in enumerate(header[29:], start=29):
                if not h: continue
                if str(h) in ('INSTRUCCIONES','ALERTAS'): break
                genericos_mb2.append(str(h)+'__MB2'); idx_mb2.append(i)

        all_gens = genericos_mb1 + genericos_mb2
        gw = {g: {} for g in all_gens}
        current_week = None

        for row in rows[4:]:
            sem = row[0]; concepto = row[2]; venta_id = str(row[1]).strip() if row[1] else ''
            if sem and str(sem).startswith('W') and concepto == 'Stock de cuartos':
                current_week = str(sem)
                for g, idx in zip(genericos_mb1, idx_mb1):
                    v = row[idx] if idx < len(row) else None
                    if isinstance(v,(int,float)) and v != 0:
                        gw[g].setdefault(current_week, {'p':0,'e':[]})['p'] += round(float(v))
                for g, idx in zip(genericos_mb2, idx_mb2):
                    v = row[idx] if idx < len(row) else None
                    if isinstance(v,(int,float)) and v != 0:
                        gw[g].setdefault(current_week, {'p':0,'e':[]})['p'] += round(float(v))
            elif concepto == 'Produccion' and current_week:
                for g, idx in zip(genericos_mb1, idx_mb1):
                    v = row[idx] if idx < len(row) else None
                    if isinstance(v,(int,float)) and v != 0:
                        gw[g].setdefault(current_week, {'p':0,'e':[]})['p'] += round(float(v))
                for g, idx in zip(genericos_mb2, idx_mb2):
                    v = row[idx] if idx < len(row) else None
                    if isinstance(v,(int,float)) and v != 0:
                        gw[g].setdefault(current_week, {'p':0,'e':[]})['p'] += round(float(v))
            elif str(concepto) not in SKIP and venta_id and current_week:
                is_res = venta_id.startswith('R-')
                is_pool = venta_id in POOL_IDS
                if not is_res and not is_pool:
                    try: int(venta_id); venta_solapa[venta_id] = solapa
                    except ValueError: pass
                pais = ''; cliente = ''
                if '\u2502' in str(concepto or ''):
                    parts = [p.strip() for p in str(concepto).split('\u2502')]
                    pais = parts[1].upper() if len(parts)>1 else ''
                    cliente = parts[2][:30] if len(parts)>2 else ''
                elif is_pool:
                    pais = venta_id.upper(); cliente = venta_id
                tipo = 'R' if is_res else ('C' if is_pool else 'V')

                for g, mb1_idx in zip(genericos_mb1, idx_mb1):
                    v = row[mb1_idx] if mb1_idx < len(row) else None
                    if isinstance(v,(int,float)) and v != 0:
                        kg = round(abs(float(v)) * KGU.get(g, 1.0))
                        f = POOL_FLAG.get(pais, FLAG_MAP.get(pais,''))
                        wkd = gw[g].setdefault(current_week, {'p':0,'e':[]})
                        found = next((e for e in wkd['e'] if e['p']==pais and e['t']==tipo and e['c']==cliente), None)
                        if found: found['kg'] += kg
                        else: wkd['e'].append({'p':pais,'c':cliente,'t':tipo,'kg':kg,'f':f})

                if is_rlw:
                    for g, mb2_idx in zip(genericos_mb2, idx_mb2):
                        v = row[mb2_idx] if mb2_idx < len(row) else None
                        if isinstance(v,(int,float)) and v != 0:
                            gen_base = g.replace('__MB2','')
                            kg = round(abs(float(v)) * KGU.get(gen_base, 1.0))
                            f = POOL_FLAG.get(pais, FLAG_MAP.get(pais,''))
                            wkd = gw[g].setdefault(current_week, {'p':0,'e':[]})
                            found = next((e for e in wkd['e'] if e['p']==pais and e['t']==tipo and e['c']==cliente), None)
                            if found: found['kg'] += kg
                            else: wkd['e'].append({'p':pais,'c':cliente,'t':tipo,'kg':kg,'f':f})

        sol_gen = {}
        for g in all_gens:
            gen_base = g.replace('__MB2','')
            kgu = KGU.get(gen_base, 1.0)
            last_v = None; week_out = {}
            for wk in sorted(gw[g].keys()):
                wn = int(wk[1:])
                if wn > 44: continue
                wd = gw[g][wk]
                week_out[wk] = {'p': round(wd['p'] * kgu), 'e': wd['e']}
                if any(e['t'] in ('V','C','R') for e in wd['e']): last_v = wk
            if week_out:
                sol_gen[g] = {'w': week_out, 'lv': last_v, 'kgu': kgu, 'mb': 'MB2' if '__MB2' in g else 'MB1'}
        sol_out[solapa] = {'g': sol_gen, 'go': [g for g in all_gens if g in sol_gen]}

    # Clientes
    cli_indiv = defaultdict(lambda: {'pais':'','flag':'','kg':0,'solapas':defaultdict(lambda: defaultdict(lambda: {'kg':0,'sem':{},'temps':set()}))})
    cong_pais = defaultdict(lambda: {'flag':'','kg':0,'clientes':set(),'solapas':defaultdict(lambda: defaultdict(lambda: {'kg':0,'sem':{}}))})

    for row in rows_dv[3:]:
        wc = row[0]; venta = str(row[5]).strip() if row[5] else ''
        pais = str(row[6]).upper() if row[6] else ''
        cliente = str(row[7]) if row[7] else ''
        generico = str(row[10]) if row[10] else ''
        tipo_prod = str(row[12]) if row[12] else ''
        kgv = row[16] if isinstance(row[16],(int,float)) else 0
        if not cliente or cliente in ('None','(en blanco)'): continue
        if not wc or not str(wc).startswith('W'): continue
        if venta.startswith('R-'): continue
        if not tipo_prod or tipo_prod == 'None': continue
        if kgv <= 0: continue
        wn = str(int(str(wc).replace('W','')))
        gen = generico if generico and generico != 'None' else '?'
        flag = FLAG_MAP.get(pais,'')
        gen_sol = venta_solapa.get(venta)
        if not gen_sol: continue

        if tipo_prod == 'Congelado':
            cp = cong_pais[pais]
            cp['flag'] = flag; cp['kg'] += kgv; cp['clientes'].add(cliente)
            gd = cp['solapas'][gen_sol][gen]
            gd['kg'] += kgv; gd['sem'][wn] = gd['sem'].get(wn,0) + kgv
        else:
            d = cli_indiv[cliente]
            if not d['pais'] and pais: d['pais']=pais; d['flag']=flag
            d['kg'] += kgv
            gd = d['solapas'][gen_sol][gen]
            gd['kg'] += kgv; gd['sem'][wn] = gd['sem'].get(wn,0) + kgv
            gd['temps'].add(tipo_prod)

    cli_out = {}
    for cli, d in sorted(cli_indiv.items(), key=lambda x:-x[1]['kg']):
        if d['kg'] < 500: continue
        solapas_out = {}
        for sol, gens in d['solapas'].items():
            gens_list = [{'g':gen,'kg':round(gd['kg']),'sem':{k:round(v) for k,v in sorted(gd['sem'].items(),key=lambda x:int(x[0]))},'temp':' + '.join(sorted(gd['temps'])) if gd['temps'] else ''} for gen,gd in sorted(gens.items(),key=lambda x:-x[1]['kg']) if gd['kg']>0]
            if gens_list:
                solapas_out[sol] = {'kg':round(sum(g['kg'] for g in gens_list)),'gens':gens_list}
        if solapas_out:
            cli_out[cli] = {'pais':d['pais'],'flag':d['flag'],'kg':round(d['kg']),'tipo':'enfriado','solapas':solapas_out}

    for pais_key, cp in sorted(cong_pais.items(), key=lambda x:-x[1]['kg']):
        if cp['kg'] < 500: continue
        nombre = f"{pais_key.title()} Congelado"
        solapas_out = {}
        for sol, gens in cp['solapas'].items():
            gens_list = [{'g':gen,'kg':round(gd['kg']),'sem':{k:round(v) for k,v in sorted(gd['sem'].items(),key=lambda x:int(x[0]))},'temp':'Congelado'} for gen,gd in sorted(gens.items(),key=lambda x:-x[1]['kg']) if gd['kg']>0]
            if gens_list:
                solapas_out[sol] = {'kg':round(sum(g['kg'] for g in gens_list)),'gens':gens_list}
        if solapas_out:
            cli_out[nombre] = {'pais':pais_key,'flag':cp['flag'],'kg':round(cp['kg']),'tipo':'congelado','clientes':sorted(cp['clientes']),'solapas':solapas_out}

    cli_out = dict(sorted(cli_out.items(), key=lambda x:-x[1]['kg']))
    result = {'sol': sol_out, 'cli': cli_out}
    s = json.dumps(result, ensure_ascii=False)
    return s.replace('&amp;', '&')


@app.route('/')
def index():
    return send_from_directory('static', 'index.html')

@app.route('/api/data')
def get_data():
    if os.path.exists(DATA_FILE):
        return app.response_class(
            open(DATA_FILE, encoding='utf-8').read(),
            mimetype='application/json'
        )
    return jsonify({'error': 'Sin datos. Subi el Excel en /admin'}), 404

@app.route('/admin', methods=['GET', 'POST'])
def admin():
    error = ''
    success = ''

    if request.args.get('logout'):
        session.clear()
        return redirect('/admin')

    if request.method == 'POST':
        if 'password' in request.form:
            if request.form['password'] == ADMIN_PASSWORD:
                session['auth'] = True
                return redirect('/admin')
            error = 'Contrasena incorrecta'
        elif session.get('auth') and 'excel' in request.files:
            f = request.files['excel']
            if f.filename.endswith('.xlsx'):
                path = os.path.join(UPLOAD_FOLDER, 'latest.xlsx')
                f.save(path)
                try:
                    data = process_excel(path)
                    open(DATA_FILE, 'w', encoding='utf-8').write(data)
                    success = 'Excel procesado correctamente. Dashboard actualizado.'
                except Exception as e:
                    error = f'Error procesando Excel: {str(e)}'
            else:
                error = 'Solo se aceptan archivos .xlsx'

    if not session.get('auth'):
        return f'''<!DOCTYPE html><html><head><meta charset="UTF-8"><title>ANB Admin</title>
<style>
body{{font-family:sans-serif;background:#0a0a0a;color:#fff;display:flex;align-items:center;justify-content:center;height:100vh;margin:0}}
.box{{background:#111;border:1px solid #c8a84b;border-radius:12px;padding:40px;min-width:320px;text-align:center}}
h2{{color:#c8a84b;margin-bottom:6px;letter-spacing:3px;font-size:16px}}
p{{color:#555;font-size:11px;margin-bottom:24px;letter-spacing:1px}}
input{{width:100%;padding:10px;background:#1a1a1a;border:1px solid #333;color:#fff;border-radius:6px;margin-bottom:12px;font-size:14px;box-sizing:border-box}}
button{{width:100%;padding:10px;background:#c8a84b;color:#000;border:none;border-radius:6px;font-weight:700;cursor:pointer;font-size:14px}}
button:hover{{background:#e0c060}}
.err{{color:#c0392b;font-size:12px;margin-bottom:10px}}
</style></head>
<body><div class="box">
<h2>ANB PLANIFICACION</h2>
<p>ACCESO ADMINISTRADOR</p>
<form method="POST">
{"<p class=err>"+error+"</p>" if error else ""}
<input type="password" name="password" placeholder="Contrasena" autofocus>
<button type="submit">Ingresar</button>
</form>
</div></body></html>'''

    return f'''<!DOCTYPE html><html><head><meta charset="UTF-8"><title>ANB Admin</title>
<style>
body{{font-family:sans-serif;background:#0a0a0a;color:#fff;display:flex;align-items:center;justify-content:center;height:100vh;margin:0}}
.box{{background:#111;border:1px solid #c8a84b;border-radius:12px;padding:40px;min-width:380px;text-align:center}}
h2{{color:#c8a84b;margin-bottom:6px;letter-spacing:3px;font-size:16px}}
p{{color:#888;font-size:12px;margin-bottom:24px}}
.upload{{border:2px dashed #333;border-radius:8px;padding:40px 20px;margin-bottom:16px;cursor:pointer;transition:border-color .2s}}
.upload:hover{{border-color:#c8a84b}}
.upload-label{{color:#c8a84b;font-size:14px;cursor:pointer}}
.upload-hint{{color:#555;font-size:11px;margin-top:6px}}
input[type=file]{{display:none}}
.btn{{display:block;width:100%;padding:10px;background:#c8a84b;color:#000;border:none;border-radius:6px;font-weight:700;cursor:pointer;font-size:14px;text-decoration:none;box-sizing:border-box}}
.btn:hover{{background:#e0c060}}
.btn-sec{{background:none;border:1px solid #333;color:#555;margin-top:8px}}
.btn-sec:hover{{border-color:#c8a84b;color:#c8a84b;background:none}}
.err{{color:#c0392b;font-size:12px;margin-bottom:12px;padding:8px;background:rgba(192,57,43,.1);border-radius:4px}}
.ok{{color:#27ae60;font-size:12px;margin-bottom:12px;padding:8px;background:rgba(39,174,96,.1);border-radius:4px}}
.links{{margin-top:20px;display:flex;gap:10px}}
</style></head>
<body><div class="box">
<h2>ACTUALIZAR DATOS</h2>
<p>Subi el Excel diario para actualizar el dashboard</p>
<form method="POST" enctype="multipart/form-data" id="form">
{"<p class=err>"+error+"</p>" if error else ""}
{"<p class=ok>"+success+"</p>" if success else ""}
<div class="upload" onclick="document.getElementById('f').click()">
<div class="upload-label">Seleccionar Excel</div>
<div class="upload-hint">.xlsx — ANB_Planificacion_Optimizada.xlsx</div>
<input type="file" id="f" name="excel" accept=".xlsx" onchange="document.getElementById('form').submit()">
</div>
</form>
<div class="links">
<a href="/" class="btn">Ver Dashboard</a>
<a href="/admin?logout=1" class="btn btn-sec">Salir</a>
</div>
</div></body></html>'''


if __name__ == '__main__':
    port = int(os.environ.get('PORT', 5000))
    app.run(host='0.0.0.0', port=port, debug=False)
